# XLSXDataSource.py
# Abstraction of an excel sheet
import openpyxl as xl
from pathlib import Path


class XLSXDataSource:
    """
    Grabs an excel workbook and handles movement through that workbook
    """

    def __init__(self, wb: xl.Workbook):
        self.wb: xl.Workbook = wb
        self.set_active_worksheet(0)

    def set_active_worksheet(self, id: str | int) -> xl.worksheet._read_only.ReadOnlyWorksheet | None:
        self.active_worksheet = None
        self.min_row = None
        self.max_row = None
        self.min_column = None
        self.max_column = None
        self.column = None
        self.row = None

        if isinstance(id, str):
            for worksheet in self.wb.worksheets:
                if worksheet.title == id:
                    self.active_worksheet = worksheet
                    break
        elif isinstance(id, int):
            if id >= 0 and id < len(self.wb.worksheets):
                self.active_worksheet = self.wb.worksheets[id]

        if self.active_worksheet is not None:
            self.min_row = self.row = self.active_worksheet.min_row
            self.max_row = self.active_worksheet.max_row
            self.min_column = self.column = self.active_worksheet.min_column
            self.max_column = self.active_worksheet.max_column

        return self.active_worksheet

    def cell_stripper(self, row, column):
        val = self.active_worksheet.cell(row, column).value
        return val.strip() if isinstance(val, str) else val

    def get_rel_cell_value(self, delta_row, delta_column):
        return self.cell_stripper(self.row + delta_row, self.column + delta_column)

    @property
    def cell(self) -> str:
        return self.cell_stripper(self.row, self.column)

    @property
    def cell_right(self) -> str:
        return self.get_rel_cell_value(0, 1)

    def move_down_one(self) -> None:
        self.row = self.row + 1

    @classmethod
    def from_file(cls, filename: Path | str):
        return cls(xl.load_workbook(filename=filename, read_only=True, data_only=True))
